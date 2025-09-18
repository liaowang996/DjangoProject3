import os
import psutil
import glob
from common import config
from common.log_utils import logger

# 兼容xls和xlsx格式（根据文件后缀选择库）
try:
    import xlrd
    from xlutils.copy import copy
except ImportError:
    logger.warning("xlrd库未安装，仅支持xlsx格式")
    xlrd = None

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font
except ImportError:
    logger.warning("openpyxl库未安装，仅支持xls格式")
    load_workbook = None


class ExcelUtils():
    def __init__(self, file_path, sheet_name):
        """初始化Excel工具类（兼容xls/xlsx）"""
        self.file_path = self._validate_file_path(file_path)
        self.sheet_name = sheet_name
        self.file_type = self._get_file_type()  # 区分xls/xlsx
        self.wb = self._open_workbook()
        self.sheet = self.get_sheets()
        self.merged_cells = self.get_merged_info()  # 缓存合并单元格
        logger.info(f"Excel工具初始化完成: {os.path.basename(self.file_path)} - {self.sheet_name}")

    def _validate_file_path(self, file_path):
        """验证文件路径有效性"""
        normalized_path = os.path.abspath(file_path)
        if not os.path.exists(normalized_path):
            err_msg = f"Excel文件不存在: {normalized_path}"
            logger.error(err_msg)
            raise FileNotFoundError(err_msg)
        if not os.path.isfile(normalized_path):
            err_msg = f"路径不是文件: {normalized_path}"
            logger.error(err_msg)
            raise IsADirectoryError(err_msg)
        # 检查文件格式支持
        if not normalized_path.endswith(('.xls', '.xlsx')):
            err_msg = f"不支持的文件格式: {normalized_path}，仅支持xls/xlsx"
            logger.error(err_msg)
            raise ValueError(err_msg)
        return normalized_path

    def _get_file_type(self):
        """获取文件类型（xls/xlsx）"""
        if self.file_path.endswith('.xlsx'):
            if not load_workbook:
                raise ImportError("openpyxl库未安装，无法处理xlsx格式")
            return 'xlsx'
        else:
            if not xlrd:
                raise ImportError("xlrd库未安装，无法处理xls格式")
            return 'xls'

    def _open_workbook(self):
        """打开工作簿（根据文件类型选择库）"""
        try:
            if self.file_type == 'xlsx':
                return load_workbook(self.file_path, data_only=True)  # data_only=True：读取单元格值而非公式
            else:
                return xlrd.open_workbook(self.file_path, formatting_info=True)
        except Exception as e:
            err_msg = f"打开Excel文件失败: {str(e)}"
            logger.error(err_msg)
            raise

    def get_sheets(self):
        """获取指定工作表"""
        try:
            if self.file_type == 'xlsx':
                return self.wb[self.sheet_name]
            else:
                return self.wb.sheet_by_name(self.sheet_name)
        except Exception as e:
            err_msg = f"工作表 '{self.sheet_name}' 不存在: {str(e)}"
            logger.error(err_msg)
            raise

    def get_row_counts(self):
        """获取行数"""
        if self.file_type == 'xlsx':
            count = self.sheet.max_row
        else:
            count = self.sheet.nrows
        logger.debug(f"工作表 '{self.sheet_name}' 行数: {count}")
        return count

    def get_col_counts(self):
        """获取列数"""
        if self.file_type == 'xlsx':
            count = self.sheet.max_column
        else:
            count = self.sheet.ncols
        logger.debug(f"工作表 '{self.sheet_name}' 列数: {count}")
        return count

    def __get_cell_value(self, row_index, col_index):
        """获取单元格值（兼容xls/xlsx，注意：xlsx行号从1开始，xls从0开始）"""
        # 统一索引：转为从0开始
        if self.file_type == 'xlsx':
            row = row_index + 1
            col = col_index + 1
            cell = self.sheet.cell(row=row, column=col)
            return cell.value if cell.value is not None else ''
        else:
            if row_index < 0 or row_index >= self.get_row_counts():
                raise IndexError(f"行索引 {row_index} 超出范围")
            if col_index < 0 or col_index >= self.get_col_counts():
                raise IndexError(f"列索引 {col_index} 超出范围")
            cell_value = self.sheet.cell_value(row_index, col_index)
            # 处理xlrd的日期格式
            if self.sheet.cell_type(row_index, col_index) == xlrd.XL_CELL_DATE:
                try:
                    return xlrd.xldate_as_datetime(cell_value, 0).strftime('%Y-%m-%d %H:%M:%S')
                except:
                    return str(cell_value)
            return cell_value if cell_value is not None else ''

    def get_merged_info(self):
        """获取合并单元格信息（返回格式：[(rlow, rhigh, clow, chigh), ...]）"""
        merged_info = []
        if self.file_type == 'xlsx':
            for merged_range in self.sheet.merged_cells.ranges:
                # 转为0开始的索引（xlsx合并范围是闭区间，如A1:A2 -> rlow=0, rhigh=2, clow=0, chigh=1）
                rlow = merged_range.min_row - 1
                rhigh = merged_range.max_row
                clow = merged_range.min_col - 1
                chigh = merged_range.max_col
                merged_info.append((rlow, rhigh, clow, chigh))
        else:
            merged_info = self.sheet.merged_cells
        logger.debug(f"工作表 '{self.sheet_name}' 合并单元格数量: {len(merged_info)}")
        return merged_info

    def get_merged_cell_value_from_cell(self, row_index, col_index):
        """获取单元格值（处理合并单元格，返回合并区域左上角的值）"""
        try:
            # 先检查是否在合并区域内
            for (rlow, rhigh, clow, chigh) in self.merged_cells:
                if rlow <= row_index < rhigh and clow <= col_index < chigh:
                    logger.debug(f"单元格({row_index},{col_index})属于合并区域，取值({rlow},{clow})")
                    return self.__get_cell_value(rlow, clow)
            # 非合并单元格，直接取值
            return self.__get_cell_value(row_index, col_index)
        except IndexError as e:
            logger.error(f"获取单元格值失败: {str(e)}")
            raise

    def get_sheet_data_by_dict(self):
        """将工作表数据转为字典列表（首行作为键）"""
        try:
            all_data = []
            row_count = self.get_row_counts()
            if row_count < 2:  # 至少需要首行（键）+1行数据
                logger.warning(f"工作表 '{self.sheet_name}' 数据不足（需至少2行）")
                return all_data

            # 获取首行作为键
            header = []
            col_count = self.get_col_counts()
            for col in range(col_count):
                key = self.get_merged_cell_value_from_cell(0, col)
                header.append(key if key else f"col_{col}")  # 空键名用col_索引代替

            # 读取数据行（从第2行开始，索引1）
            for row in range(1, row_count):
                row_data = {}
                for col in range(col_count):
                    key = header[col]
                    value = self.get_merged_cell_value_from_cell(row, col)
                    row_data[key] = value
                all_data.append(row_data)

            logger.info(f"从工作表 '{self.sheet_name}' 读取到 {len(all_data)} 条数据")
            return all_data
        except Exception as e:
            logger.error(f"转换工作表数据失败: {str(e)}", exc_info=True)
            raise

    def read_multiple_excel_cases(self, sheet_name='Sheet1'):
        """批量读取指定目录下所有Excel文件的用例（修复原参数未使用问题）"""
        all_cases = []
        # 从配置获取测试数据目录（原代码用self.test_data，改为从config读取）
        excel_dir_path = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', config.TEST_DATA))

        try:
            if not os.path.exists(excel_dir_path):
                logger.error(f"测试数据目录不存在: {excel_dir_path}")
                return all_cases
            if not os.path.isdir(excel_dir_path):
                logger.error(f"路径不是目录: {excel_dir_path}")
                return all_cases

            # 查找所有Excel文件
            excel_files = glob.glob(os.path.join(excel_dir_path, '*.xls')) + \
                          glob.glob(os.path.join(excel_dir_path, '*.xlsx'))
            if not excel_files:
                logger.warning(f"目录 {excel_dir_path} 中未找到Excel文件")
                return all_cases

            logger.info(f"发现 {len(excel_files)} 个Excel文件，开始读取用例...")
            for file_path in excel_files:
                file_name = os.path.basename(file_path)
                try:
                    # 读取当前文件的指定工作表
                    excel = ExcelUtils(file_path, sheet_name)
                    cases = excel.get_sheet_data_by_dict()
                    if cases:
                        # 添加来源文件字段，便于追溯
                        for case in cases:
                            case['来源文件'] = file_name
                        all_cases.extend(cases)
                        logger.info(f"从 {file_name} 读取到 {len(cases)} 条用例")
                    else:
                        logger.warning(f"{file_name} 中未读取到用例数据")
                except Exception as e:
                    logger.error(f"处理文件 {file_name} 失败: {str(e)}", exc_info=True)
                    continue  # 继续处理下一个文件

            logger.info(f"批量读取完成，共获取 {len(all_cases)} 条用例")
            return all_cases
        except Exception as e:
            logger.error(f"批量读取Excel用例异常: {str(e)}", exc_info=True)
            return all_cases

    def _is_file_locked(self):
        """检查文件是否被占用"""
        try:
            with open(self.file_path, 'a'):
                return False
        except IOError:
            return True

    def force_close_file(self):
        """强制关闭占用文件的进程"""
        file_path = os.path.abspath(self.file_path)
        for proc in psutil.process_iter(['pid', 'name', 'open_files']):
            try:
                for file in proc.info['open_files'] or []:
                    if os.path.abspath(file.path) == file_path:
                        logger.warning(f"发现占用进程: {proc.info['name']} (PID: {proc.info['pid']})")
                        proc.terminate()
                        proc.wait(timeout=2)
                        logger.info(f"已关闭进程 {proc.info['pid']}")
                        return True
            except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess) as e:
                logger.warning(f"处理进程时出错: {str(e)}")
        return False

    def close(self):
        """关闭工作簿"""
        try:
            if self.wb:
                self.wb.close()
                logger.debug(f"已关闭Excel文件: {os.path.basename(self.file_path)}")
        except Exception as e:
            logger.warning(f"关闭Excel文件失败: {str(e)}")


if __name__ == '__main__':
    # 测试代码
    current_path = os.path.dirname(os.path.abspath(__file__))
    # 测试xlsx文件（需确保test_demo.xlsx存在）
    xlsx_path = os.path.join(current_path, '..', 'test_data', 'test_demo.xlsx')
    if os.path.exists(xlsx_path):
        excel_xlsx = ExcelUtils(xlsx_path, 'Sheet1')
        print(f"xlsx数据: {excel_xlsx.get_sheet_data_by_dict()}")
        excel_xlsx.close()
    # 测试xls文件
    xls_path = os.path.join(current_path, '..', 'test_data', 'test_demo.xls')
    if os.path.exists(xls_path):
        excel_xls = ExcelUtils(xls_path, 'Sheet1')
        print(f"xls数据: {excel_xls.get_sheet_data_by_dict()}")
        excel_xls.close()
    # 测试批量读取
    ExcelUtils(xls_path, 'Sheet1').read_multiple_excel_cases()